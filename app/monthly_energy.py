from flask import request, jsonify, Response
from datetime import datetime as dt
from io import BytesIO, StringIO
import csv
import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None


def _to_float(val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    try:
        return float(str(val).strip())
    except:
        return None


def parse_excel_records(file_bytes):
    records = []
    headers = {}

    def normalize_header(text):
        if text is None:
            return ""
        return str(text).replace(" ", "").replace("\n", "").strip()

    rows = []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        rows = [tuple(cell for cell in row) for row in wb.active.iter_rows(values_only=True)]
    except Exception:
        if xlrd is not None:
            try:
                book = xlrd.open_workbook(file_contents=file_bytes)
                sheet = book.sheet_by_index(0)
                rows = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]
            except Exception:
                rows = []
        if not rows:
            try:
                text = file_bytes.decode('utf-8-sig')
            except Exception:
                text = file_bytes.decode('gbk', errors='replace')
            rows = [tuple(r) for r in csv.reader(StringIO(text))]

    for row in rows:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        header_texts = [normalize_header(cell) for cell in row if isinstance(cell, str)]
        row_is_header = any('名称' in text for text in header_texts) and any('地点' in text for text in header_texts)
        if row_is_header:
            headers = {}
            for idx, cell in enumerate(row):
                val = normalize_header(cell)
                if not val:
                    continue
                if "月份" in val or "report_month" in val or "reportmonth" in val:
                    headers[idx] = "report_month"
                elif "名称" in val or ("名" in val and "倍率" not in val and "上月" not in val):
                    headers[idx] = "name"
                elif "等级" in val:
                    headers[idx] = "level"
                elif "地点" in val:
                    headers[idx] = "location"
                elif "本月读数" in val:
                    headers[idx] = "current_reading"
                elif "上月读数" in val:
                    headers[idx] = "last_reading"
                elif "倍率" in val or val == "倍数":
                    headers[idx] = "multiplier"
                elif "实际用电量" in val or "实际用量" in val or "实际差值" in val:
                    headers[idx] = "actual_usage"
                elif "上月用电" in val or "用水量" in val:
                    headers[idx] = "prev_month_usage"
            continue

        if not headers:
            continue

        rec = {
            "report_month": None,
            "name": None,
            "level": None,
            "location": None,
            "last_reading": None,
            "current_reading": None,
            "multiplier": None,
            "actual_usage": None,
            "prev_month_usage": None,
        }
        for idx, cell in enumerate(row):
            col = headers.get(idx)
            if not col:
                continue
            if col in {"name", "level", "location", "report_month"}:
                rec[col] = str(cell).strip() if cell is not None else None
            else:
                rec[col] = _to_float(cell)

        if not rec["name"] and rec["location"]:
            rec["name"] = rec["location"]
        if not rec["name"]:
            continue

        if rec["actual_usage"] is None and rec["last_reading"] is not None and rec["current_reading"] is not None:
            multiplier = rec["multiplier"] if rec["multiplier"] is not None else 1
            rec["actual_usage"] = (rec["current_reading"] - rec["last_reading"] ) * multiplier

        if all(rec[k] is None for k in ["last_reading", "current_reading", "multiplier", "actual_usage", "prev_month_usage"]):
            continue
        records.append(rec)

    return records


def parse_template_records(file_bytes):
    """Parse a template file containing 名称 / 等级 / 地点 columns and return list of {name, level, location}."""
    rows = []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        rows = [tuple(cell for cell in row) for row in wb.active.iter_rows(values_only=True)]
    except Exception:
        if xlrd is not None:
            try:
                book = xlrd.open_workbook(file_contents=file_bytes)
                sheet = book.sheet_by_index(0)
                rows = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]
            except Exception:
                rows = []
        if not rows:
            try:
                text = file_bytes.decode('utf-8-sig')
            except Exception:
                text = file_bytes.decode('gbk', errors='replace')
            rows = [tuple(r) for r in csv.reader(StringIO(text))]

    def normalize(x):
        return str(x).replace(' ', '').replace('\n', '').strip() if x is not None else ''

    headers = {}
    out = []
    for row in rows:
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        # detect header row
        htexts = [normalize(c) for c in row if isinstance(c, str) or c is not None]
        if any('名称' in t for t in htexts) or any('名' in t for t in htexts):
            headers = {}
            for idx, cell in enumerate(row):
                val = normalize(cell)
                if not val:
                    continue
                if '名称' in val or (('名' in val) and '倍率' not in val and '上月' not in val):
                    headers[idx] = 'name'
                elif '等级' in val:
                    headers[idx] = 'level'
                elif '地点' in val:
                    headers[idx] = 'location'
            continue
        if not headers:
            continue
        rec = {'name': None, 'level': None, 'location': None}
        for idx, cell in enumerate(row):
            col = headers.get(idx)
            if not col:
                continue
            rec[col] = str(cell).strip() if cell is not None else None
        if rec['name']:
            out.append(rec)
    return out


def register_monthly_energy_routes(app, get_db, require_admin):
    @app.route("/api/monthly_energy_entries")
    def list_monthly_energy_entries():
        month = request.args.get("report_month", "")
        conn = get_db()
        q = "SELECT * FROM monthly_energy_entries WHERE 1=1"
        params = []
        if month:
            q += " AND report_month=?"
            params.append(month)
        rows = conn.execute(q + " ORDER BY report_month DESC, name ASC", params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])

    @app.route("/api/monthly_energy_entries", methods=["POST"])
    def save_monthly_energy_entry():
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        d = request.json or {}
        report_month = (d.get("report_month") or dt.now().strftime("%Y-%m")).strip()
        name = (d.get("name") or "").strip()
        if not name:
            return jsonify({"error": "名称不能为空"}), 400
        conn = get_db()
        conn.execute("""INSERT INTO monthly_energy_entries
            (report_month,name,level,location,last_reading,current_reading,multiplier,actual_usage,prev_month_usage,notes,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(report_month,name,location) DO UPDATE SET
            level=excluded.level, location=excluded.location,
            last_reading=excluded.last_reading, current_reading=excluded.current_reading,
            multiplier=excluded.multiplier, actual_usage=excluded.actual_usage,
            prev_month_usage=excluded.prev_month_usage, notes=excluded.notes,
            updated_at=excluded.updated_at""",
            (report_month, name, d.get("level"), d.get("location"), d.get("last_reading"),
             d.get("current_reading"), d.get("multiplier"), d.get("actual_usage"),
             d.get("prev_month_usage"), d.get("notes",""), dt.now().isoformat()))
        conn.commit()
        row = conn.execute("SELECT id FROM monthly_energy_entries WHERE report_month=? AND name=? AND location=?",
                           (report_month, name, d.get("location"))).fetchone()
        conn.close()
        return jsonify({"ok": True, "id": row["id"] if row else None})

    @app.route("/api/monthly_energy_entries/<int:mid>", methods=["PUT"])
    def update_monthly_energy_entry(mid):
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        d = request.json or {}
        conn = get_db()
        row = conn.execute("SELECT * FROM monthly_energy_entries WHERE id=?", (mid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "不存在"}), 404
        report_month = (d.get("report_month") or row["report_month"]).strip()
        name = (d.get("name") or row["name"]).strip()
        if not name:
            conn.close()
            return jsonify({"error": "名称不能为空"}), 400
        conn.execute("""UPDATE monthly_energy_entries SET
            report_month=?, name=?, level=?, location=?, last_reading=?, current_reading=?,
            multiplier=?, actual_usage=?, prev_month_usage=?, notes=?, updated_at=?
            WHERE id=?""",
            (report_month, name, d.get("level", row["level"]), d.get("location", row["location"]),
             d.get("last_reading", row["last_reading"]), d.get("current_reading", row["current_reading"]),
             d.get("multiplier", row["multiplier"]), d.get("actual_usage", row["actual_usage"]),
             d.get("prev_month_usage", row["prev_month_usage"]), d.get("notes", row["notes"]),
             dt.now().isoformat(), mid))
        conn.commit(); conn.close()
        return jsonify({"ok": True})

    @app.route("/api/monthly_energy_entries/<int:mid>", methods=["DELETE"])
    def delete_monthly_energy_entry(mid):
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        conn = get_db()
        conn.execute("DELETE FROM monthly_energy_entries WHERE id=?", (mid,))
        conn.commit(); conn.close()
        return jsonify({"ok": True})

    @app.route("/api/monthly_energy/import", methods=["POST"])
    def import_monthly_energy():
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        if "file" not in request.files:
            return jsonify({"error": "请上传 Excel 文件"}), 400
        default_month = (request.form.get("report_month") or dt.now().strftime("%Y-%m")).strip()
        f = request.files["file"]
        try:
            records = parse_excel_records(f.read())
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        if not records:
            return jsonify({"error": "未识别到可导入的记录"}), 400
        conn = get_db(); imported = 0
        for rec in records:
            try:
                row_month = (rec.get("report_month") or default_month or dt.now().strftime("%Y-%m")).strip()
                actual_usage = rec.get("actual_usage")
                if actual_usage is None and rec.get("last_reading") is not None and rec.get("current_reading") is not None:
                    multiplier = rec.get("multiplier") if rec.get("multiplier") is not None else 1
                    actual_usage = (rec["current_reading"] - rec["last_reading"]) * multiplier
                conn.execute("""INSERT INTO monthly_energy_entries
                    (report_month,name,level,location,last_reading,current_reading,multiplier,actual_usage,prev_month_usage,notes,updated_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(report_month,name,location) DO UPDATE SET
                    level=excluded.level, location=excluded.location,
                    last_reading=excluded.last_reading, current_reading=excluded.current_reading,
                    multiplier=excluded.multiplier, actual_usage=excluded.actual_usage,
                    prev_month_usage=excluded.prev_month_usage, notes=excluded.notes,
                    updated_at=excluded.updated_at""",
                    (row_month, rec["name"], rec["level"], rec["location"], rec["last_reading"],
                     rec["current_reading"], rec["multiplier"], actual_usage,
                     rec["prev_month_usage"], "", dt.now().isoformat()))
                imported += 1
            except:
                pass
        conn.commit(); conn.close()
        return jsonify({"ok": True, "report_month": default_month, "imported": imported, "rows": len(records)})

    @app.route("/api/monthly_energy/export")
    def export_monthly_energy():
        report_month = request.args.get("report_month", "")
        fmt = request.args.get("format", "xlsx").lower()
        # support template export via existing export endpoint for robustness
        want_template = request.args.get('template') in ('1', 'true', 'yes')
        if want_template:
            # generate same template as export_monthly_template
            conn = get_db()
            rows = conn.execute("SELECT name, level, location FROM monthly_energy_templates ORDER BY name ASC").fetchall()
            conn.close()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Template'
            ws.append(["月份","名称","等级","地点","上月读数","本月读数","倍率","实际用电量","上月用电","备注"])
            if rows:
                for r in rows:
                    ws.append(["", r["name"] or "", r["level"] or "", r["location"] or "", "", "", "", "", "", ""])
            else:
                ws.append(["","示例名称","示例等级","示例地点","","","","","",""])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment;filename=monthly_template.xlsx"})
        # continue normal export
        conn = get_db()
        q = "SELECT * FROM monthly_energy_entries WHERE 1=1"
        params = []
        if report_month:
            q += " AND report_month=?"
            params.append(report_month)
        rows = conn.execute(q + " ORDER BY report_month DESC, name ASC", params).fetchall()
        conn.close()

        if fmt == 'csv':
            lines = ["\ufeff月份,名称,等级,地点,上月读数,本月读数,倍率,实际用电量,上月用电,备注"]
            for r in rows:
                lines.append(",".join([
                    str(r["report_month"] or ""),
                    str(r["name"] or ""),
                    str(r["level"] or ""),
                    str(r["location"] or ""),
                    str(r["last_reading"] or ""),
                    str(r["current_reading"] or ""),
                    str(r["multiplier"] or ""),
                    str(r["actual_usage"] or ""),
                    str(r["prev_month_usage"] or ""),
                    str(r["notes"] or "")
                ]))
            fname = f"monthly_energy_{report_month or 'all'}.csv"
            return Response("\n".join(lines), mimetype="text/csv",
                headers={"Content-Disposition": f"attachment;filename={fname}"})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'MonthlyData'
        ws.append(["月份","名称","等级","地点","上月读数","本月读数","倍率","实际用电量","上月用电","备注"])
        for r in rows:
            ws.append([
                r["report_month"] or "",
                r["name"] or "",
                r["level"] or "",
                r["location"] or "",
                r["last_reading"] if r["last_reading"] is not None else "",
                r["current_reading"] if r["current_reading"] is not None else "",
                r["multiplier"] if r["multiplier"] is not None else "",
                r["actual_usage"] if r["actual_usage"] is not None else "",
                r["prev_month_usage"] if r["prev_month_usage"] is not None else "",
                r["notes"] or ""
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f"monthly_energy_{report_month or 'all'}.xlsx"
        return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={fname}"})

    @app.route("/api/monthly_energy/template/export")
    def export_monthly_template():
        # Export a minimal template containing only 名称 / 等级 / 地点
        conn = get_db()
        rows = conn.execute("SELECT name, level, location FROM monthly_energy_templates ORDER BY name ASC").fetchall()
        conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        ws.append(["名称", "等级", "地点"])
        if rows:
            for r in rows:
                ws.append([r["name"] or "", r["level"] or "", r["location"] or ""])
        else:
            ws.append(["示例名称", "示例等级", "示例地点"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=monthly_template.xlsx"})

    # Alias route to support environments where the longer path may 404
    @app.route("/api/monthly_template.xlsx")
    def export_monthly_template_alias():
        return export_monthly_template()

    @app.route("/api/monthly_energy_templates")
    def list_monthly_energy_templates():
        conn = get_db()
        rows = conn.execute("SELECT * FROM monthly_energy_templates ORDER BY name ASC").fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])

    @app.route("/api/monthly_energy/template/import", methods=["POST"])
    def import_monthly_template():
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        if "file" not in request.files:
            return jsonify({"error": "请上传模板文件"}), 400
        f = request.files["file"]
        try:
            recs = parse_template_records(f.read())
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        if not recs:
            return jsonify({"error": "未识别到模板记录"}), 400
        conn = get_db(); saved = 0
        for r in recs:
            try:
                name = (r.get('name') or '').strip()
                if not name:
                    continue
                conn.execute("INSERT INTO monthly_energy_templates(name,level,location,notes,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(name,location) DO UPDATE SET level=excluded.level, location=excluded.location, notes=excluded.notes, updated_at=excluded.updated_at",
                             (name, r.get('level'), r.get('location'), '', dt.now().isoformat()))
                saved += 1
            except:
                pass
        conn.commit(); conn.close()
        return jsonify({"ok": True, "saved": saved, "rows": len(recs)})

    @app.route("/api/monthly_energy_templates", methods=["POST"])
    def save_monthly_energy_template():
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        d = request.json or {}
        name = (d.get("name") or "").strip()
        if not name:
            return jsonify({"error": "名称不能为空"}), 400
        conn = get_db()
        conn.execute("INSERT INTO monthly_energy_templates(name,level,location,notes,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(name,location) DO UPDATE SET level=excluded.level, location=excluded.location, notes=excluded.notes, updated_at=excluded.updated_at",
                     (name, d.get("level"), d.get("location"), d.get("notes",""), dt.now().isoformat()))
        conn.commit()
        row = conn.execute("SELECT id FROM monthly_energy_templates WHERE name=? AND location=?", (name, d.get("location"))).fetchone()
        conn.close()
        return jsonify({"ok": True, "id": row["id"] if row else None})

    @app.route("/api/monthly_energy_templates/<int:tid>", methods=["DELETE"])
    def delete_monthly_energy_template(tid):
        if not require_admin(request):
            return jsonify({"error": "权限不足"}), 403
        conn = get_db()
        conn.execute("DELETE FROM monthly_energy_templates WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    return None