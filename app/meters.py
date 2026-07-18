"""
分表抄表台账（meter_readings）模块
----------------------------------
独立于 server.py 的 Flask Blueprint，包含：
- 数据库表初始化（init_db，由 server.py 的 init_db() 调用）
- 不规则多分区抄表 Excel 的解析（_parse_meter_excel）
- 抄表条目的增删改查 / Excel 导入导出 API

通过在各函数内延迟导入 server 来复用 get_db() / get_session()，
避免循环导入问题。
"""

from flask import Blueprint, request, jsonify, Response
from datetime import datetime as dt
import csv
import io

try:
    import xlrd
except ImportError:
    xlrd = None

meters_bp = Blueprint("meters", __name__)


def init_db(conn):
    """由 server.py 的 init_db() 在建库时调用，创建本模块自己的表"""
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS meter_readings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period TEXT NOT NULL,
        category TEXT DEFAULT '',
        meter_type TEXT DEFAULT '电',
        name TEXT NOT NULL,
        grade TEXT DEFAULT '',
        location TEXT DEFAULT '',
        prev_reading REAL,
        curr_reading REAL,
        multiplier REAL DEFAULT 1,
        usage_amount REAL DEFAULT 0,
        notes TEXT DEFAULT '',
        sort_order INTEGER DEFAULT 0,
        created_by TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_meter_period ON meter_readings(period);

    CREATE TABLE IF NOT EXISTS billing_area_readings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period TEXT NOT NULL,
        category TEXT DEFAULT "",
        meter_type TEXT DEFAULT "",
        name TEXT NOT NULL,
        grade TEXT DEFAULT "",
        location TEXT DEFAULT "",
        prev_reading REAL,
        curr_reading REAL,
        multiplier REAL DEFAULT 1,
        usage_amount REAL DEFAULT 0,
        notes TEXT DEFAULT "",
        sort_order INTEGER DEFAULT 0,
        created_by TEXT DEFAULT "",
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_billing_period ON billing_area_readings(period);

    CREATE TABLE IF NOT EXISTS cost_comparison (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        period TEXT NOT NULL,
        energy_type TEXT NOT NULL,
        current_year REAL,
        current_price REAL,
        current_amount REAL,
        previous_year REAL,
        previous_price REAL,
        previous_amount REAL,
        comparison_rate REAL,
        unit TEXT DEFAULT '',
        created_by TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_cost_period ON cost_comparison(period);
    """)


def require_login_write(req):
    """抄表台账专用：无论是否有 token，一律要求已登录且非 viewer"""
    import server
    token = req.headers.get("X-Session-Token", "")
    if not token:
        print('[meters] require_login_write: missing token')
        return False
    sess = server.get_session(token)
    if not sess:
        print('[meters] require_login_write: invalid token', token)
        return False
    print('[meters] require_login_write: token ok', token, sess)
    return sess.get("role") != "viewer"


# ══════════════════════════════════════════════════════════
# 分表抄表台账（meter_readings）
# ══════════════════════════════════════════════════════════

def _meter_row_usage(prev, curr, mult):
    if prev is None or curr is None or mult is None:
        return None
    try:
        d = (float(curr) - float(prev)) * float(mult)
        return round(d, 2) if d >= 0 else None
    except (TypeError, ValueError):
        return None

@meters_bp.route("/api/meters/periods")
def meter_periods():
    import server
    conn = server.get_db()
    rows = conn.execute("SELECT DISTINCT period FROM meter_readings ORDER BY period DESC").fetchall()
    conn.close()
    return jsonify([r["period"] for r in rows])

@meters_bp.route("/api/meters")
def list_meters():
    import server
    period = request.args.get("period", "")
    conn = server.get_db()
    if period:
        rows = conn.execute(
            "SELECT * FROM meter_readings WHERE period=? ORDER BY sort_order ASC, id ASC", (period,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM meter_readings ORDER BY period DESC, sort_order ASC, id ASC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@meters_bp.route("/api/meters/summary")
def meter_summary():
    import server
    period = request.args.get("period", "")
    conn = server.get_db()
    q = "SELECT meter_type, COUNT(*) c, SUM(usage_amount) s FROM meter_readings WHERE 1=1"
    p = []
    if period:
        q += " AND period=?"; p.append(period)
    q += " GROUP BY meter_type"
    rows = conn.execute(q, p).fetchall()
    q2 = "SELECT COUNT(DISTINCT category) c FROM meter_readings WHERE 1=1"
    if period: q2 += " AND period=?"
    cat_count = conn.execute(q2, p).fetchone()["c"]
    conn.close()
    by_type = {r["meter_type"]: {"count": r["c"], "total": round(r["s"] or 0, 2)} for r in rows}
    return jsonify({"period": period, "by_type": by_type, "category_count": cat_count})


@meters_bp.route("/api/meters/floor-summary")
def meter_floor_summary():
    """按楼层统计用电量，根据location字段提取楼层，排除走道"""
    import server
    period = request.args.get("period", "")
    conn = server.get_db()
    
    # 获取所有数据（排除走道和空地点的大厦总表）
    q = "SELECT name, location, meter_type, usage_amount FROM meter_readings WHERE 1=1"
    p = []
    if period:
        q += " AND period=?"; p.append(period)
    q += " AND location NOT LIKE '%走道%' AND location IS NOT NULL AND trim(location) != ''"
    rows = conn.execute(q, p).fetchall()
    conn.close()
    
    # 按楼层分组统计
    import re
    floor_stats = {}
    # 按名称独立统计的地点（不按楼层归类，直接以名称作为条目）
    independent_names = {"消控室", "消控中心", "酷迪咖啡", "库迪咖啡", "库底咖啡"}
    # 按名称排除的地点（不参与统计）
    excluded_names = {}
    # 特殊地点映射到楼层（这些地点没有楼层号但属于特定楼层）
    special_location_map = {
        "打卡处": "-1F",
        "大堂": "1F",
        "副井道": "2F",
        "技术层大厅": "2F",
        "技术层弱电间": "5F",
        "技术层大厅西": "5F",
        "电梯机房": "5F",
        "经济房地下库": "-1F",
        "卫星机房": "-1F",
        "立体车库": "-1F",
        "锅炉室": "-1F",
        "经济房地下": "-1F",
        "弱电间": "-1F",
        "安利仓库": "-1F",
    }

    for row in rows:
        name = row["name"] or ""
        location = row["location"] or ""
        meter_type = row["meter_type"] or "电"
        usage = row["usage_amount"] or 0

        # 排除不需要统计的名称
        if name in excluded_names:
            continue

        # 按名称独立统计的地点
        if name in independent_names:
            floor = name
        elif name in special_location_map:
            floor = special_location_map[name]
        elif location in special_location_map:
            floor = special_location_map[location]
        else:
            # 提取楼层：支持 "—1F电工房"、"1F强电间"、"爱康1F"、"16楼强电井" 等
            # 用 search 匹配字符串中任意位置的楼层模式（支持 F 和 楼）
            floor_match = re.search(r'(\d+)(F|楼)', location)
            if floor_match:
                raw = floor_match.group(1) + "F"  # 统一为 "1F" 格式
                # 检查楼层数字前面是否有破折号（中文—或英文-）
                before = location[:floor_match.start()]
                floor = "-" + raw if before and before[-1] in ("—", "-") else raw
            else:
                floor = "其他"

        if floor not in floor_stats:
            floor_stats[floor] = {"电": 0, "水": 0}

        if meter_type == "水":
            floor_stats[floor]["水"] += usage
        else:
            floor_stats[floor]["电"] += usage
    
    # 按楼层排序（-1F, 1F, 2F, ...），独立统计地点排在最后
    def floor_sort_key(f):
        if f in independent_names:
            return (1, f)  # 独立统计地点排在后面
        m = re.match(r'(-?\d+)F', f)
        if m:
            return (0, int(m.group(1)))
        return (2, f)  # "其他" 排在最后
    
    sorted_floors = sorted(floor_stats.keys(), key=floor_sort_key)
    
    result = []
    for floor in sorted_floors:
        result.append({
            "floor": floor,
            "电": floor_stats[floor]["电"],
            "水": floor_stats[floor]["水"]
        })
    
    return jsonify(result)

@meters_bp.route("/api/meters", methods=["POST"])
def add_meter():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再录入"}), 403
    d = request.json or {}
    period = (d.get("period") or "").strip()
    name = (d.get("name") or "").strip()
    if not period or not name:
        return jsonify({"error": "月份和名称必填"}), 400
    prev, curr, mult = d.get("prev_reading"), d.get("curr_reading"), d.get("multiplier", 1)
    usage = d.get("usage_amount")
    if usage is None:
        usage = _meter_row_usage(prev, curr, mult) or 0
    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn = server.get_db()
    conn.execute("""INSERT INTO meter_readings
        (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,notes,created_by,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (period, d.get("category", ""), d.get("meter_type", "电"), name, d.get("grade", ""),
         d.get("location", ""), prev, curr, mult, usage, d.get("notes", ""),
         sess["username"] if sess else "", dt.now().isoformat()))
    conn.commit()
    mid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "id": mid})

@meters_bp.route("/api/meters/<int:mid>", methods=["PUT"])
def update_meter(mid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再修改"}), 403
    d = request.json or {}
    conn = server.get_db()
    row = conn.execute("SELECT * FROM meter_readings WHERE id=?", (mid,)).fetchone()
    if not row: conn.close(); return jsonify({"error": "记录不存在"}), 404
    def pick(k, fb):
        v = d.get(k, None)
        return v if v is not None else fb
    prev = pick("prev_reading", row["prev_reading"])
    curr = pick("curr_reading", row["curr_reading"])
    mult = pick("multiplier", row["multiplier"])
    usage = d.get("usage_amount")
    if usage is None:
        calc = _meter_row_usage(prev, curr, mult)
        usage = calc if calc is not None else row["usage_amount"]
    conn.execute("""UPDATE meter_readings SET period=?,category=?,meter_type=?,name=?,grade=?,
        location=?,prev_reading=?,curr_reading=?,multiplier=?,usage_amount=?,notes=?,updated_at=? WHERE id=?""",
        (pick("period", row["period"]), pick("category", row["category"]), pick("meter_type", row["meter_type"]),
         pick("name", row["name"]), pick("grade", row["grade"]), pick("location", row["location"]),
         prev, curr, mult, usage, pick("notes", row["notes"]), dt.now().isoformat(), mid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@meters_bp.route("/api/meters/<int:mid>", methods=["DELETE"])
def del_meter(mid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再删除"}), 403
    conn = server.get_db()
    conn.execute("DELETE FROM meter_readings WHERE id=?", (mid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@meters_bp.route("/api/meters/batch-delete", methods=["POST"])
def batch_delete_meters():
    import server
    if not require_login_write(request):
        return jsonify({"error": "请先登录后台账号再删除"}), 403
    data = request.get_json(silent=True) or {}
    ids = data.get("ids")
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "请提供要删除的记录ID列表"}), 400
    # 过滤并转换为整数 ID 列表
    ids = [int(i) for i in ids if isinstance(i, (int, str)) and str(i).isdigit()]
    if not ids:
        return jsonify({"error": "要删除的记录ID列表不能为空"}), 400

    try:
        conn = server.get_db()
        deleted = 0
        # SQLite 默认对变量数量有限制（通常为 999），对大量 ID 分批执行删除
        chunk_size = 500
        for i in range(0, len(ids), chunk_size):
            chunk = ids[i:i+chunk_size]
            placeholders = ",".join(["?"] * len(chunk))
            before = conn.total_changes
            conn.execute(f"DELETE FROM meter_readings WHERE id IN ({placeholders})", chunk)
            after = conn.total_changes
            deleted += (after - before)
        conn.commit()
        # total_changes is cumulative since connection opened; return min(len(ids), deleted)
        conn.close()
        return jsonify({"ok": True, "deleted": len(ids)})
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        print('[meters] batch_delete_meters error', repr(e))
        return jsonify({"error": str(e)}), 500


def _norm(s):
    return s.replace(" ", "").replace("\u3000", "").replace("\xa0", "").lower()

import json

def _normalize_header(text):
    if text is None:
        return ""
    t = str(text).strip().lower()
    for ch in [" ", "\u3000", "\xa0", "\n", "\r", "\t", ":", "：", "(", ")", "[", "]"]:
        t = t.replace(ch, "")
    return t

def _guess_header_key(text):
    norm = _normalize_header(text)
    if not norm:
        return None
    mapping = {
        "名称": "name", "名称/设备": "name", "设备名称": "name", "表名": "name", "户名": "name", "编号": "name", "机组": "name", "表号": "name", "设备": "name", "登记": "name", "登记号": "name", "区域名称": "name",
        "等级": "grade", "级别": "grade", "类别": "grade", "规格": "grade",
        "地点": "location", "位置": "location", "安装地点": "location", "房间": "location", "区域": "location", "部门": "location",
        "上月读数": "prev", "上月份读数": "prev", "上月表数": "prev", "前月读数": "prev", "上期读数": "prev", "上月读数kw.h": "prev", "上月读数kwh": "prev",
        "本月读数": "curr", "本月表数": "curr", "本期读数": "curr", "本月读数kw.h": "curr", "本月读数kwh": "curr",
        "倍率": "mult", "倍数": "mult", "乘数": "mult",
        "实际用电量": "usage", "实际用量": "usage", "实际差值": "usage", "本月实际数": "usage", "实际用电量kw.h": "usage", "实际用量kw.h": "usage", "用电量": "usage", "用量": "usage",
        "分类": "category", "分区": "category", "区域": "category", "分区/区域": "category",
        "表类型": "meter_type", "类型": "meter_type"
    }
    for key, value in mapping.items():
        if key in norm:
            return value
    return None

def _find_header_row(rows):
    for ri, row in enumerate(rows):
        guessed = {}
        for i, v in enumerate(row):
            key = _guess_header_key(v)
            if key:
                guessed[i] = key
        if guessed and ("name" in guessed.values() or "prev" in guessed.values() or "curr" in guessed.values() or "usage" in guessed.values()):
            return ri, guessed
    return None, None


def _validate_electricity_stats(wb, filename=""):
    """验证文件是否为大厦用电统计表
    
    扫描所有 sheet 的前若干行，查找标题行：
    「浙江出版物资大厦  XXXX年X月份用电统计表」（日期可变）
    找到则通过，否则拒绝导入。
    返回 (is_valid: bool, error_message: str or None)
    """
    import re
    fname = (filename or "").lower()
    if "抄表" in fname:
        return False, "该文件疑似抄表数据，请前往「收费区域」导入"

    pattern = re.compile(r"浙江出版物资大厦.*\d{4}\s*年\s*\d{1,2}\s*月份用电统计表")

    for ws in wb.worksheets:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 20:
                break
            for cell in row:
                if cell is not None and pattern.search(str(cell)):
                    return True, None
            row_text = "".join(str(v) for v in row if v is not None)
            if pattern.search(row_text):
                return True, None

    return False, "未识别到「浙江出版物资大厦 XXXX年X月份用电统计表」标题，请确认文件是否为用电统计表"


def _load_workbook(file_bytes, filename=""):
    """将 .xls / .xlsx / .csv 文件统一转换为 openpyxl Workbook
    
    返回 (openpyxl.Workbook, error_msg_or_None)
    """
    import openpyxl
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""

    # --- 尝试 openpyxl（.xlsx） ---
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        return wb, None
    except Exception:
        pass

    # --- 尝试 xlrd（.xls） ---
    if xlrd is not None:
        try:
            book = xlrd.open_workbook(file_contents=file_bytes)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sname in book.sheet_names():
                ws = wb.create_sheet(title=sname)
                sheet = book.sheet_by_name(sname)
                for r in range(sheet.nrows):
                    row_vals = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    ws.append(row_vals)
            return wb, None
        except Exception:
            pass

    # --- 尝试 CSV ---
    try:
        try:
            text = file_bytes.decode("utf-8-sig")
        except Exception:
            text = file_bytes.decode("gbk", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for row in rows:
            ws.append(row)
        return wb, None
    except Exception:
        pass

    return None, "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"


def _validate_billing_file(wb):
    """验证文件是否为收费区域抄表数据文件，检查是否包含关键字"""
    required_keywords = ("水电抄表", "水抄表", "电抄表")
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    for keyword in required_keywords:
                        if keyword in cell:
                            return True
    return False


def _parse_meter_excel(wb, custom_mapping=None, header_row_idx=None, period=None, sheets=None):
    """把不规则的多分区抄表 Excel 解析成扁平的条目列表
    
    参数:
        wb: openpyxl Workbook
        custom_mapping: 自定义列映射
        header_row_idx: 表头行索引
        period: 所属月份（YYYY-MM），如果传入则覆盖自动检测
        sheets: 指定要解析的工作表列表（ws 对象），为 None 则解析所有工作表
    """
    SKIP_NAME_MARKERS = ("合计", "总表", "用电量", "用水量", "实际用电", "物业部", "工程部")
    entries = []
    sort_order = 0
    worksheets = sheets if sheets is not None else wb.worksheets
    
    def _extract_period_from_text(text):
        if not text:
            return None
        import re
        # 严格匹配日期格式，避免误匹配读数等数字
        patterns = [
            re.compile(r"(\d{4})\s*[\-/\.年]\s*(\d{1,2})\s*月?"),
            re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月"),
        ]
        for p in patterns:
            m = p.search(text)
            if m:
                try:
                    y = int(m.group(1))
                    mo = int(m.group(2))
                    if 1900 <= y <= 2100 and 1 <= mo <= 12:
                        return f"{y:04d}-{mo:02d}"
                except:
                    continue
        return None
    
    def _is_header_row(vals):
        header_keywords = ['名称', '地点', '读数', '倍率', '用量']
        found = 0
        for v in vals:
            if v:
                for kw in header_keywords:
                    if kw in v:
                        found += 1
                        break
        return found >= 2

    for ws in worksheets:
        category = ws.title if ws.title and not ws.title.lower().startswith("sheet") else ""
        meter_type = "电"
        header_map = None
        current_period = period
        if custom_mapping is not None:
            header_map = {int(col): field for field, col in custom_mapping.items() if col is not None}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [(str(v).strip() if v is not None else "") for v in row]
            non_empty = [(i, v) for i, v in enumerate(vals) if v != ""]
            if not non_empty:
                continue
            joined = _normalize_header("".join(v for _, v in non_empty))

            if custom_mapping is not None and header_row_idx is not None and row_idx <= header_row_idx:
                continue

            first_val = non_empty[0][1] if non_empty else ""
            
            if "水电抄表" in first_val or "抄表单" in first_val:
                extracted = _extract_period_from_text(first_val)
                if extracted:
                    current_period = extracted
                # 同时从这类行中提取category名称（如"3F黄庭水、电抄表单"→"3F黄庭"）
                raw = str(first_val).strip()
                for suffix in ["水、电抄表单", "水电抄表单", "抄表单", "水、电抄表", "水电抄表"]:
                    if raw.endswith(suffix):
                        raw = raw[:-len(suffix)].rstrip("：: ")
                        break
                if raw:
                    category = raw
                continue

            if len(non_empty) == 1:
                text = _normalize_header(non_empty[0][1]).rstrip(":：")
                
                # 跳过纯数字行（如 46133.0）
                try:
                    float(text)
                    continue
                except (ValueError, TypeError):
                    pass
                
                if text in ("电表", "水表"):
                    meter_type = "水" if "水" in non_empty[0][1] else "电"
                    if custom_mapping is None:
                        header_map = None
                    continue
                
                if "电表" in text or "水表" in text:
                    meter_type = "水" if "水" in text else "电"
                    if custom_mapping is None:
                        header_map = None
                    continue
                
                if any(m in text for m in ("合计", "总表", "物业部", "工程部")):
                    continue
                
                if "水" in text and "表" in text and "抄表" not in text:
                    meter_type = "水"
                # 清理category名称，去掉"水、电抄表单"等后缀
                category = text.rstrip("：:")
                for suffix in ["水、电抄表单", "水电抄表单", "抄表单", "水、电抄表", "水电抄表"]:
                    if category.endswith(suffix) or category.endswith(suffix.lower()):
                        matched = suffix if category.endswith(suffix) else suffix.lower()
                        category = category[:-len(matched)].rstrip("：: ")
                        break
                continue

            # 检查是否是区域汇总行（用电量: X, 用水量: Y）跨多列
            # 例如: {0: '用电量:', 1: '17546.0', 2: '用水量:', 3: '978.0'}
            if len(non_empty) >= 2 and category:
                import re
                joined_text = " ".join(v for _, v in non_empty)
                elec_match = re.search(r'用电量[:：]\s*([\d.]+)', joined_text)
                water_match = re.search(r'[冷热]?用水量[:：]\s*([\d.]+)', joined_text)
                if elec_match:
                    elec_usage = float(elec_match.group(1))
                    water_usage = float(water_match.group(1)) if water_match else 0
                    
                    sort_order += 1
                    entries.append({
                        "period": current_period,
                        "category": category,
                        "meter_type": "电",
                        "name": category + "总用电",
                        "grade": "",
                        "location": "",
                        "prev_reading": None,
                        "curr_reading": None,
                        "multiplier": 1,
                        "usage_amount": elec_usage,
                        "sort_order": sort_order,
                    })
                    
                    if water_usage > 0:
                        sort_order += 1
                        entries.append({
                            "period": current_period,
                            "category": category,
                            "meter_type": "水",
                            "name": category + "总用水",
                            "grade": "",
                            "location": "",
                            "prev_reading": None,
                            "curr_reading": None,
                            "multiplier": 1,
                            "usage_amount": water_usage,
                            "sort_order": sort_order,
                        })
                    continue

            if custom_mapping is None:
                if _is_header_row(vals):
                    guessed = {}
                    for i, v in non_empty:
                        key = _guess_header_key(v)
                        if key:
                            guessed[i] = key
                    if guessed and ("name" in guessed.values() or "prev" in guessed.values() or "curr" in guessed.values() or "usage" in guessed.values()):
                        if header_map is not None:
                            old_map = header_map.copy()
                            old_map.update(guessed)
                            header_map = old_map
                        else:
                            header_map = guessed
                    continue

            if header_map:
                rec = {}
                for i, v in enumerate(vals):
                    key = header_map.get(i)
                    if key:
                        rec[key] = v
                name = rec.get("name", "").strip()
                if not name or any(m in name for m in SKIP_NAME_MARKERS):
                    continue
                def num(x):
                    try:
                        return float(str(x).replace(',', '').replace('，', ''))
                    except (TypeError, ValueError):
                        return None
                prev = num(rec.get("prev"))
                curr = num(rec.get("curr"))
                mult = num(rec.get("mult"))
                usage = num(rec.get("usage"))
                if usage is None:
                    usage = _meter_row_usage(prev, curr, mult if mult is not None else 1)
                if prev is None and curr is None and usage is None:
                    continue
                sort_order += 1
                entries.append({
                    "period": current_period,
                    "category": rec.get("category", category),
                    "meter_type": meter_type,
                    "name": name,
                    "grade": rec.get("grade", ""),
                    "location": rec.get("location", ""),
                    "prev_reading": prev,
                    "curr_reading": curr,
                    "multiplier": mult if mult is not None else (1 if usage is not None else None),
                    "usage_amount": usage or 0,
                    "sort_order": sort_order,
                })
    return entries


def _guess_period_from_wb(wb):
    """尝试从工作簿的表名或前几行单元格中识别出月份，返回格式 YYYY-MM 或 None
    
    支持多种格式：2026年7月、2026/7、2026-07、202607、7月、七月等
    当只识别到月份时，使用当前年份。
    """
    import re
    # 严格匹配日期格式，避免误匹配读数等数字
    # 要求年份和月份之间有明确的分隔符（年、-、/、.）或"月"字
    patterns = [
        re.compile(r"(\d{4})\s*[\-/\.年]\s*(\d{1,2})\s*月?"),
        re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月"),
    ]
    month_only_pattern = re.compile(r"(\d{1,2})\s*月")
    chinese_month_pattern = re.compile(r"(一|二|三|四|五|六|七|八|九|十|十一|十二)\s*月")
    chinese_to_num = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6,
                      "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11, "十二": 12}
    
    def norm_year_month(m):
        try:
            y = int(m[0]); mo = int(m[1])
            if 1900 <= y <= 2100 and 1 <= mo <= 12:
                return f"{y:04d}-{mo:02d}"
        except Exception:
            return None
        return None

    def find_with_year(text):
        for p in patterns:
            m = p.search(text)
            if m:
                res = norm_year_month(m.groups())
                if res:
                    return res
        return None

    def find_month_only(text):
        m = month_only_pattern.search(text)
        if m:
            try:
                mo = int(m.group(1))
                if 1 <= mo <= 12:
                    return mo
            except Exception:
                pass
        m = chinese_month_pattern.search(text)
        if m:
            mo = chinese_to_num.get(m.group(1))
            if mo:
                return mo
        return None

    # 先检查每个 sheet 的前几行和前几列的单元格文本（带年份）
    for ws in wb.worksheets:
        try:
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx > 20:
                    break
                for col_idx, cell in enumerate(row[:6]):
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    result = find_with_year(text)
                    if result:
                        return result
        except Exception:
            continue

    # 如果单元格中没找到，再检查每个 sheet 的标题（带年份）
    for ws in wb.worksheets:
        title = (ws.title or "").strip()
        result = find_with_year(title)
        if result:
            return result

    # 如果没找到带年份的，先从单元格识别月份
    for ws in wb.worksheets:
        try:
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx > 20:
                    break
                for col_idx, cell in enumerate(row[:6]):
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    month = find_month_only(text)
                    if month:
                        return f"{dt.now().year}-{month:02d}"
        except Exception:
            continue

    # 如果单元格中没找到，再从 sheet 标题识别月份
    for ws in wb.worksheets:
        title = (ws.title or "").strip()
        month = find_month_only(title)
        if month:
            return f"{dt.now().year}-{month:02d}"

    return None


def _detect_sheet_period(ws):
    """检测单个工作表的所属月份（YYYY-MM）
    
    优先从前几行单元格内容中识别抄表日期，其次从 sheet 标题中识别。
    支持多种格式：2026年7月、2026/7、2026-07、202607、7月、七月等
    当只识别到月份时，使用当前年份。
    """
    import re
    # 严格匹配日期格式，避免误匹配读数等数字
    # 要求年份和月份之间有明确的分隔符（年、-、/、.）或"月"字
    patterns = [
        re.compile(r"(\d{4})\s*[\-/\.年]\s*(\d{1,2})\s*月?"),
        re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月"),
    ]
    month_only_pattern = re.compile(r"(\d{1,2})\s*月")
    chinese_month_pattern = re.compile(r"(一|二|三|四|五|六|七|八|九|十|十一|十二)\s*月")
    chinese_to_num = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6,
                      "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11, "十二": 12}
    
    def norm_year_month(m):
        try:
            y = int(m[0]); mo = int(m[1])
            if 1900 <= y <= 2100 and 1 <= mo <= 12:
                return f"{y:04d}-{mo:02d}"
        except Exception:
            return None
        return None

    # 先找带年份的模式
    def find_with_year(text):
        for p in patterns:
            m = p.search(text)
            if m:
                res = norm_year_month(m.groups())
                if res:
                    return res
        return None

    # 先检查前 10 行的单元格，优先从表格内容中识别日期（检查所有列）
    try:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 10:
                break
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                result = find_with_year(text)
                if result:
                    return result
    except Exception:
        pass

    # 如果表格内容中没找到日期，再检查 sheet 标题
    title = (ws.title or "").strip()
    result = find_with_year(title)
    if result:
        return result

    # 如果没找到带年份的，尝试只识别月份
    def find_month_only(text):
        m = month_only_pattern.search(text)
        if m:
            try:
                mo = int(m.group(1))
                if 1 <= mo <= 12:
                    return mo
            except Exception:
                pass
        m = chinese_month_pattern.search(text)
        if m:
            mo = chinese_to_num.get(m.group(1))
            if mo:
                return mo
        return None

    # 先检查单元格中的月份
    try:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 10:
                break
            for cell in row[:6]:
                if cell is None:
                    continue
                text = str(cell).strip()
                month = find_month_only(text)
                if month:
                    return f"{dt.now().year}-{month:02d}"
    except Exception:
        pass

    # 如果单元格中没找到，再从 sheet 标题识别月份
    month = find_month_only(title)
    if month:
        return f"{dt.now().year}-{month:02d}"

    return None

@meters_bp.route("/api/meters/import", methods=["POST"])
def import_meters():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再导入"}), 403
    period = (request.form.get("period") or "").strip()
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择要导入的 Excel 文件"}), 400

    mapping_json = request.form.get("mapping")
    header_row_idx = request.form.get("header_row_idx")
    file_content = file.read()
    try:
        wb, err = _load_workbook(file_content, file.filename or "")
        if wb is None:
            return jsonify({"error": err or "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {e}"}), 400

    is_valid, content_err = _validate_electricity_stats(wb, file.filename or "")
    if not is_valid:
        return jsonify({"error": content_err}), 400

    custom_mapping = None
    if mapping_json:
        try:
            parsed_mapping = json.loads(mapping_json)
            if isinstance(parsed_mapping, dict):
                custom_mapping = {}
                for field, col in parsed_mapping.items():
                    if col is None or col == "" or (isinstance(col, str) and not col.strip().isdigit()):
                        custom_mapping[field] = None
                    else:
                        try:
                            custom_mapping[field] = int(col)
                        except Exception:
                            custom_mapping[field] = None
                if not any(v is not None for v in custom_mapping.values()):
                    custom_mapping = None
        except Exception:
            custom_mapping = None
    if header_row_idx is not None:
        try:
            header_row_idx = int(header_row_idx)
        except Exception:
            header_row_idx = None

    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn = server.get_db()

    # ── 模式 1: 前端显式指定了月份 ──
    # 只解析第一个工作表（最符合用户预期）
    if period:
        entries = _parse_meter_excel(wb, custom_mapping=custom_mapping, header_row_idx=header_row_idx, period=period, sheets=[wb.worksheets[0]])
        if not entries:
            conn.close()
            return jsonify({"error": f"月份 {period} 的工作表中未能识别出有效数据行，请检查格式后再导入"}), 400

        existing = conn.execute("SELECT COUNT(*) FROM meter_readings WHERE period=?", (period,)).fetchone()[0]
        conn.execute("DELETE FROM meter_readings WHERE period=?", (period,))
        username = sess["username"] if sess else ""
        now = dt.now().isoformat()
        rows = [(period, e["category"], e["meter_type"], e["name"], e["grade"], e["location"],
                 e["prev_reading"], e["curr_reading"], e["multiplier"], e["usage_amount"], e["sort_order"],
                 username, now) for e in entries]
        conn.executemany("""INSERT INTO meter_readings
            (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,sort_order,created_by,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
        conn.commit(); conn.close()
        result = {"ok": True, "imported": len(entries), "period": period, "matched_sheets": len(matched_sheets)}
        if existing > 0:
            result["overwritten"] = existing
        return jsonify(result)

    # ── 模式 2: 自动分月份模式 ──
    # 按 sheet 分别识别月份，已存在的月份跳过，只导入新月份
    # 关键：每个 sheet 独立 auto-detect 表头，不传入 custom_mapping
    # 以避免不同 sheet 结构差异导致的数据丢失
    total_imported = 0
    total_skipped = 0
    sheet_results = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        sheet_period = _detect_sheet_period(ws)
        if not sheet_period:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": None,
                "status": "skipped",
                "reason": "未能自动识别月份",
                "count": 0
            })
            total_skipped += 1
            continue

        # 检查该月份是否已存在
        existing = conn.execute(
            "SELECT COUNT(*) FROM meter_readings WHERE period=?", (sheet_period,)
        ).fetchone()[0]

        if existing > 0:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": sheet_period,
                "status": "skipped",
                "reason": f"月份 {sheet_period} 已有 {existing} 条数据，自动跳过",
                "count": existing
            })
            total_skipped += 1
            continue

        # 解析该 sheet 的数据
        # 不传 custom_mapping 和 header_row_idx，让每个 sheet 独立 auto-detect
        sheet_entries = _parse_meter_excel(
            wb,
            period=sheet_period,
            sheets=[ws]
        )

        if not sheet_entries:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": sheet_period,
                "status": "skipped",
                "reason": "该 sheet 未能识别出有效数据行",
                "count": 0
            })
            total_skipped += 1
            continue

        # 插入数据
        for e in sheet_entries:
            conn.execute("""INSERT INTO meter_readings
                (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,sort_order,created_by,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (sheet_period, e["category"], e["meter_type"], e["name"], e["grade"], e["location"],
                 e["prev_reading"], e["curr_reading"], e["multiplier"], e["usage_amount"], e["sort_order"],
                 sess["username"] if sess else "", dt.now().isoformat()))

        total_imported += len(sheet_entries)
        sheet_results.append({
            "sheet_name": sheet_name,
            "period": sheet_period,
            "status": "imported",
            "count": len(sheet_entries)
        })

    conn.commit(); conn.close()

    if total_imported == 0:
        return jsonify({
            "error": "所有月份的数据已存在或无法识别，无需导入",
            "sheet_results": sheet_results,
            "total_imported": 0,
            "total_skipped": total_skipped
        }), 400

    return jsonify({
        "ok": True,
        "imported": total_imported,
        "skipped": total_skipped,
        "sheet_results": sheet_results
    })

@meters_bp.route("/api/meters/import_preview", methods=["POST"])
def import_preview():
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择要解析的 Excel 文件"}), 400

    try:
        wb, err = _load_workbook(file.read(), file.filename or "")
        if wb is None:
            return jsonify({"error": err or "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {e}"}), 400

    is_valid, content_err = _validate_electricity_stats(wb, file.filename or "")
    if not is_valid:
        return jsonify({"error": content_err}), 400

    raw_headers = []
    raw_rows = []
    header_row_idx = None
    auto_map = {}

    field_map = {}
    for ws in wb.worksheets:
        rows = [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        idx, guessed = _find_header_row(rows)
        if idx is not None:
            header_row_idx = idx
            raw_headers = rows[header_row_idx]
            raw_rows = rows[header_row_idx + 1:header_row_idx + 11]
            field_map = {field: col for col, field in guessed.items()}
            break

    if header_row_idx is None:
        for ws in wb.worksheets:
            rows = [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]
            for idx, row in enumerate(rows):
                if any(cell != "" for cell in row):
                    header_row_idx = idx
                    raw_headers = rows[header_row_idx]
                    raw_rows = rows[header_row_idx + 1:header_row_idx + 11]
                    break
            if header_row_idx is not None:
                break

    parsed_rows = _parse_meter_excel(wb, custom_mapping=field_map or None, header_row_idx=header_row_idx) if header_row_idx is not None else _parse_meter_excel(wb)

    # 尝试从工作簿中识别所属月份（YYYY-MM）
    detected_period = None
    try:
        detected_period = _guess_period_from_wb(wb)
    except Exception:
        detected_period = None

    # 按 sheet 检测月份信息
    sheet_periods = []
    for ws in wb.worksheets:
        sp = _detect_sheet_period(ws)
        sheet_periods.append({
            "sheet_name": ws.title,
            "period": sp or None
        })

    sample_parsed = []
    for entry in parsed_rows[:10]:
        sample_parsed.append({
            "name": entry.get("name", ""),
            "grade": entry.get("grade", ""),
            "location": entry.get("location", ""),
            "prev_reading": entry.get("prev_reading"),
            "curr_reading": entry.get("curr_reading"),
            "multiplier": entry.get("multiplier"),
            "usage_amount": entry.get("usage_amount"),
        })

    # 检查已存在哪些月份
    import server as srv
    exists_map = {}
    try:
        conn_exists = srv.get_db()
        for sp in sheet_periods:
            if sp["period"]:
                cnt = conn_exists.execute("SELECT COUNT(*) FROM meter_readings WHERE period=?", (sp["period"],)).fetchone()[0]
                exists_map[sp["period"]] = cnt
        conn_exists.close()
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "header_row": header_row_idx if header_row_idx is not None else 0,
        "headers": raw_headers,
        "raw_rows": raw_rows,
        "parsed_rows": sample_parsed,
        "detected_period": detected_period,
        "sheet_periods": sheet_periods,
        "exists_map": exists_map,
        "mapping": field_map,
    })

@meters_bp.route("/api/meters/export")
def export_meters():
    import server
    period = request.args.get("period", "")
    fmt = request.args.get("format", "xlsx").lower()
    conn = server.get_db()
    if period:
        rows = conn.execute(
            "SELECT * FROM meter_readings WHERE period=? ORDER BY category, sort_order, id", (period,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM meter_readings ORDER BY period DESC, category, sort_order, id").fetchall()
    conn.close()

    # 格式化月份为用户友好的显示
    def _fmt_period(p):
        try:
            y, m = p.split("-")
            return f"{int(y)}年{int(m)}月"
        except Exception:
            return p

    title = _fmt_period(period) + " 抄表台账" if period else "全部月份 抄表台账"

    if fmt == "csv":
        lines = ["\ufeff" + title]
        lines.append("名称,等级,地点,上月读数 KW.h,本月读数 KW.h,倍率,实际用电量 KW.h,月份")
        for r in rows:
            lines.append(",".join([
                str(r["name"] or ""),
                str(r["grade"] or ""),
                str(r["location"] or ""),
                str(r["prev_reading"] if r["prev_reading"] is not None else ""),
                str(r["curr_reading"] if r["curr_reading"] is not None else ""),
                str(r["multiplier"] if r["multiplier"] is not None else ""),
                str(r["usage_amount"] if r["usage_amount"] is not None else ""),
                str(r["period"] or ""),
            ]))
        fname = f"抄表台账_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
        from urllib.parse import quote
        fname_encoded = quote(fname.encode("utf-8"), safe="")
        return Response("\n".join(lines), mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})

    # 默认 xlsx
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "抄表台账"

    # 标题行
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 表头行
    headers = ["名称", "等级", "地点", "上月读数 KW.h", "本月读数 KW.h", "倍率", "实际用电量 KW.h"]
    if not period:
        headers.append("月份")
    ws.append(headers)
    header_row_idx = 2
    for c in ws[header_row_idx]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="DDEBF7")
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        row_data = [r["name"], r["grade"], r["location"],
                     r["prev_reading"], r["curr_reading"], r["multiplier"], r["usage_amount"]]
        if not period:
            row_data.append(r["period"])
        ws.append(row_data)

    widths = [24, 10, 16, 16, 16, 10, 18]
    if not period:
        widths.append(12)
    for i, w in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"抄表台账_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from urllib.parse import quote
    fname_encoded = quote(fname.encode("utf-8"), safe="")
    return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})

@meters_bp.route("/api/meters/template")
def template_meters():
    """导出空白抄表模板"""
    import openpyxl, io
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "抄表台账"
    
    # 设置表头
    headers = ["名称", "等级", "地点", "上月读数 KW.h", "本月读数 KW.h", "倍率", "实际用电量 KW.h"]
    ws.append(headers)
    
    # 设置表头样式
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="366092")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 添加示例行和说明
    example_data = [
        ["2026-07", "一楼", "电", "总表", "高压", "配电房", 1000, 1050, 1, 50, ""],
        ["2026-07", "二楼", "水", "冷却水表", "", "冷却塔", 5000, 5100, 1, 100, ""],
        ["2026-07", "三楼", "气", "燃气表", "", "锅炉房", 500, 550, 1, 50, ""],
    ]
    
    for row_data in example_data:
        ws.append(row_data)
    
    # 清空示例数据，保留格式（用户可删除）
    for row in ws.iter_rows(min_row=2, max_row=4):
        for cell in row:
            cell.value = None
    
    # 设置列宽和对齐
    widths = [12, 16, 8, 16, 10, 16, 14, 14, 8, 12, 20]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i)
        ws.column_dimensions[col_letter].width = w
    
    # 为所有行添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # 非表头行
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加说明页
    ws_guide = wb.create_sheet("使用说明")
    guide_text = [
        ["抄表台账导入模板使用说明"],
        [""],
        ["1. 列说明:"],
        ["   - 月份: 必填，格式为 YYYY-MM，如 2026-07"],
        ["   - 分区/区域: 必填，如一楼、二楼、锅炉房等"],
        ["   - 类型: 必填，可选值为 电、水、气"],
        ["   - 名称: 必填，表名称"],
        ["   - 等级: 选填，如高压、低压、冷却等"],
        ["   - 地点: 选填，表位置描述"],
        ["   - 上月读数: 选填，数值类型"],
        ["   - 本月读数: 必填（若要自动计算用量），数值类型"],
        ["   - 倍率: 选填，默认为 1"],
        ["   - 实际用量: 选填，若为空则自动计算为 (本月-上月)×倍率"],
        ["   - 备注: 选填"],
        [""],
        ["2. 操作步骤:"],
        ["   ① 下载模板"],
        ["   ② 在 '抄表台账' 表单中填写或修改数据"],
        ["   ③ 保存文件"],
        ["   ④ 在后台管理中选择月份后上传该文件"],
        ["   ⑤ 数据将覆盖该月份的所有现有数据"],
        [""],
        ["3. 数据格式要求:"],
        ["   - 请勿修改列顺序"],
        ["   - 请勿删除表头行"],
        ["   - 数值类型字段请输入数字（可以是小数）"],
        ["   - 不要在表头行下方插入空行"],
    ]
    
    for row_data in guide_text:
        ws_guide.append(row_data)
    
    # 设置说明页的格式
    ws_guide.column_dimensions['A'].width = 50
    for row in ws_guide.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"抄表台账模板_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # 处理中文文件名，使用 RFC 5987 编码
    from urllib.parse import quote
    fname_encoded = quote(fname.encode('utf-8'), safe='')
    headers_dict = {"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"}
    return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_dict)


# =============================================================================
# 收费区域（billing_area_readings）
# =============================================================================

@meters_bp.route("/api/billing/periods")
def billing_periods():
    import server
    conn = server.get_db()
    rows = conn.execute("SELECT DISTINCT period FROM billing_area_readings ORDER BY period DESC").fetchall()
    conn.close()
    return jsonify([r["period"] for r in rows])


@meters_bp.route("/api/billing/summary")
def billing_summary():
    import server
    try:
        period = request.args.get("period", "")
        conn = server.get_db()
        if period:
            rows = conn.execute(
                "SELECT category, meter_type, SUM(usage_amount) as total_usage, COUNT(*) as count "
                "FROM billing_area_readings WHERE period=? GROUP BY category, meter_type ORDER BY category, meter_type",
                (period,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT period, category, meter_type, SUM(usage_amount) as total_usage, COUNT(*) as count "
                "FROM billing_area_readings GROUP BY period, category, meter_type ORDER BY period DESC, category, meter_type"
            ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        print('[meters] billing_summary error:', e)
        return jsonify({"error": "查询失败: " + str(e)}), 500


@meters_bp.route("/api/billing")
def list_billing():
    import server
    try:
        period = request.args.get("period", "")
        conn = server.get_db()
        if period:
            rows = conn.execute(
                "SELECT * FROM billing_area_readings WHERE period=? ORDER BY sort_order ASC, id ASC", (period,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM billing_area_readings ORDER BY period DESC, sort_order ASC, id ASC").fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        print('[meters] list_billing error:', e)
        return jsonify({"error": "查询失败: " + str(e)}), 500


@meters_bp.route("/api/billing", methods=["POST"])
def add_billing():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再录入"}), 403
    d = request.json or {}
    period = (d.get("period") or "").strip()
    name = (d.get("name") or "").strip()
    if not period or not name:
        return jsonify({"error": "月份和名称必填"}), 400
    prev, curr, mult = d.get("prev_reading"), d.get("curr_reading"), d.get("multiplier", 1)
    usage = d.get("usage_amount")
    if usage is None:
        usage = _meter_row_usage(prev, curr, mult) or 0
    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn = server.get_db()
    conn.execute("""INSERT INTO billing_area_readings
        (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,notes,created_by,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (period, d.get("category", ""), d.get("meter_type", "电"), name, d.get("grade", ""),
         d.get("location", ""), prev, curr, mult, usage, d.get("notes", ""),
         sess["username"] if sess else "", dt.now().isoformat()))
    conn.commit()
    mid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "id": mid})


@meters_bp.route("/api/billing/<int:bid>", methods=["GET"])
def get_billing(bid):
    import server
    conn = server.get_db()
    row = conn.execute("SELECT * FROM billing_area_readings WHERE id=?", (bid,)).fetchone()
    if not row: conn.close(); return jsonify({"error": "记录不存在"}), 404
    conn.close()
    return jsonify(dict(row))


@meters_bp.route("/api/billing/<int:bid>", methods=["PUT"])
def update_billing(bid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再修改"}), 403
    d = request.json or {}
    conn = server.get_db()
    row = conn.execute("SELECT * FROM billing_area_readings WHERE id=?", (bid,)).fetchone()
    if not row: conn.close(); return jsonify({"error": "记录不存在"}), 404
    def pick(k, fb):
        v = d.get(k, None)
        return v if v is not None else fb
    prev = pick("prev_reading", row["prev_reading"])
    curr = pick("curr_reading", row["curr_reading"])
    mult = pick("multiplier", row["multiplier"])
    usage = d.get("usage_amount")
    if usage is None:
        calc = _meter_row_usage(prev, curr, mult)
        usage = calc if calc is not None else row["usage_amount"]
    conn.execute("""UPDATE billing_area_readings SET period=?,category=?,meter_type=?,name=?,grade=?,
        location=?,prev_reading=?,curr_reading=?,multiplier=?,usage_amount=?,notes=?,updated_at=? WHERE id=?""",
        (pick("period", row["period"]), pick("category", row["category"]), pick("meter_type", row["meter_type"]),
         pick("name", row["name"]), pick("grade", row["grade"]), pick("location", row["location"]),
         prev, curr, mult, usage, pick("notes", row["notes"]), dt.now().isoformat(), bid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@meters_bp.route("/api/billing/<int:bid>", methods=["DELETE"])
def del_billing(bid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再删除"}), 403
    conn = server.get_db()
    conn.execute("DELETE FROM billing_area_readings WHERE id=?", (bid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@meters_bp.route("/api/billing/batch-delete", methods=["POST"])
def batch_delete_billing():
    import server
    if not require_login_write(request):
        return jsonify({"error": "请先登录后台账号再删除"}), 403
    data = request.get_json(silent=True) or {}
    ids = data.get("ids")
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "请提供要删除的记录ID列表"}), 400
    ids = [int(i) for i in ids if isinstance(i, (int, str)) and str(i).isdigit()]
    if not ids:
        return jsonify({"error": "要删除的记录ID列表不能为空"}), 400

    try:
        conn = server.get_db()
        deleted = 0
        chunk_size = 500
        for i in range(0, len(ids), chunk_size):
            chunk = ids[i:i+chunk_size]
            placeholders = ",".join(["?"] * len(chunk))
            before = conn.total_changes
            conn.execute(f"DELETE FROM billing_area_readings WHERE id IN ({placeholders})", chunk)
            after = conn.total_changes
            deleted += (after - before)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "deleted": len(ids)})
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        return jsonify({"error": str(e)}), 500


@meters_bp.route("/api/billing/import", methods=["POST"])
def import_billing():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再导入"}), 403
    period = (request.form.get("period") or "").strip()
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择要导入的 Excel 文件"}), 400
    mapping_json = request.form.get("mapping")
    header_row_idx = request.form.get("header_row_idx")
    try:
        wb, err = _load_workbook(file.read(), file.filename or "")
        if wb is None:
            return jsonify({"error": err or "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {e}"}), 400

    if not _validate_billing_file(wb):
        return jsonify({"error": "文件内容校验失败，请上传包含「水电抄表」、「水抄表」或「电抄表」关键字的抄表数据文件"}), 400

    custom_mapping = None
    if mapping_json:
        try:
            parsed_mapping = json.loads(mapping_json)
            if isinstance(parsed_mapping, dict):
                custom_mapping = {}
                for field, col in parsed_mapping.items():
                    if col is None or col == "" or (isinstance(col, str) and not col.strip().isdigit()):
                        custom_mapping[field] = None
                    else:
                        try:
                            custom_mapping[field] = int(col)
                        except Exception:
                            custom_mapping[field] = None
                if not any(v is not None for v in custom_mapping.values()):
                    custom_mapping = None
        except Exception:
            custom_mapping = None
    if header_row_idx is not None:
        try:
            header_row_idx = int(header_row_idx)
        except Exception:
            header_row_idx = None

    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn = server.get_db()

    if period:
        entries = _parse_meter_excel(wb, custom_mapping=custom_mapping, header_row_idx=header_row_idx, period=period, sheets=[wb.worksheets[0]])
        if not entries:
            conn.close()
            return jsonify({"error": f"月份 {period} 的工作表中未能识别出有效数据行，请检查格式后再导入"}), 400

        existing = conn.execute("SELECT COUNT(*) FROM billing_area_readings WHERE period=?", (period,)).fetchone()[0]
        conn.execute("DELETE FROM billing_area_readings WHERE period=?", (period,))
        username = sess["username"] if sess else ""
        now = dt.now().isoformat()
        rows = [(period, e["category"], e["meter_type"], e["name"], e["grade"], e["location"],
                 e["prev_reading"], e["curr_reading"], e["multiplier"], e["usage_amount"], e["sort_order"],
                 username, now) for e in entries]
        conn.executemany("""INSERT INTO billing_area_readings
            (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,sort_order,created_by,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
        conn.commit(); conn.close()
        result = {"ok": True, "imported": len(entries), "period": period}
        if existing > 0:
            result["overwritten"] = existing
        return jsonify(result)

    total_imported = 0
    total_skipped = 0
    sheet_results = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        sheet_period = _detect_sheet_period(ws)
        if not sheet_period:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": None,
                "status": "skipped",
                "reason": "未能自动识别月份",
                "count": 0
            })
            total_skipped += 1
            continue

        existing = conn.execute(
            "SELECT COUNT(*) FROM billing_area_readings WHERE period=?", (sheet_period,)
        ).fetchone()[0]

        if existing > 0:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": sheet_period,
                "status": "skipped",
                "reason": f"月份 {sheet_period} 已有 {existing} 条数据，自动跳过",
                "count": existing
            })
            total_skipped += 1
            continue

        sheet_entries = _parse_meter_excel(
            wb,
            period=sheet_period,
            sheets=[ws]
        )

        if not sheet_entries:
            sheet_results.append({
                "sheet_name": sheet_name,
                "period": sheet_period,
                "status": "skipped",
                "reason": "该 sheet 未能识别出有效数据行",
                "count": 0
            })
            total_skipped += 1
            continue

        for e in sheet_entries:
            conn.execute("""INSERT INTO billing_area_readings
                (period,category,meter_type,name,grade,location,prev_reading,curr_reading,multiplier,usage_amount,sort_order,created_by,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (sheet_period, e["category"], e["meter_type"], e["name"], e["grade"], e["location"],
                 e["prev_reading"], e["curr_reading"], e["multiplier"], e["usage_amount"], e["sort_order"],
                 sess["username"] if sess else "", dt.now().isoformat()))

        total_imported += len(sheet_entries)
        sheet_results.append({
            "sheet_name": sheet_name,
            "period": sheet_period,
            "status": "imported",
            "count": len(sheet_entries)
        })

    conn.commit(); conn.close()

    if total_imported == 0:
        return jsonify({
            "error": "所有月份的数据已存在或无法识别，无需导入",
            "sheet_results": sheet_results,
            "total_imported": 0,
            "total_skipped": total_skipped
        }), 400

    return jsonify({
        "ok": True,
        "imported": total_imported,
        "skipped": total_skipped,
        "sheet_results": sheet_results
    })


@meters_bp.route("/api/billing/import_preview", methods=["POST"])
def billing_import_preview():
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请选择要解析的 Excel 文件"}), 400
    try:
        wb, err = _load_workbook(file.read(), file.filename or "")
        if wb is None:
            return jsonify({"error": err or "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {e}"}), 400

    if not _validate_billing_file(wb):
        return jsonify({"error": "文件内容校验失败，请上传包含「水电抄表」、「水抄表」或「电抄表」关键字的抄表数据文件"}), 400

    raw_headers = []
    raw_rows = []
    header_row_idx = None
    auto_map = {}

    field_map = {}
    for ws in wb.worksheets:
        rows = [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        idx, guessed = _find_header_row(rows)
        if idx is not None:
            header_row_idx = idx
            raw_headers = rows[header_row_idx]
            raw_rows = rows[header_row_idx + 1:header_row_idx + 11]
            field_map = {field: col for col, field in guessed.items()}
            break

    if header_row_idx is None:
        for ws in wb.worksheets:
            rows = [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            for idx, row in enumerate(rows):
                if any(cell != "" for cell in row):
                    header_row_idx = idx
                    raw_headers = rows[header_row_idx]
                    raw_rows = rows[header_row_idx + 1:header_row_idx + 11]
                    break
            if header_row_idx is not None:
                break

    parsed_rows = _parse_meter_excel(wb, custom_mapping=field_map or None, header_row_idx=header_row_idx) if header_row_idx is not None else _parse_meter_excel(wb)

    detected_period = None
    try:
        detected_period = _guess_period_from_wb(wb)
    except Exception:
        detected_period = None

    sheet_periods = []
    for ws in wb.worksheets:
        sp = _detect_sheet_period(ws)
        sheet_periods.append({
            "sheet_name": ws.title,
            "period": sp or None
        })

    sample_parsed = []
    for entry in parsed_rows[:10]:
        sample_parsed.append({
            "name": entry.get("name", ""),
            "grade": entry.get("grade", ""),
            "location": entry.get("location", ""),
            "prev_reading": entry.get("prev_reading"),
            "curr_reading": entry.get("curr_reading"),
            "multiplier": entry.get("multiplier"),
            "usage_amount": entry.get("usage_amount"),
        })

    import server as srv
    exists_map = {}
    try:
        conn_exists = srv.get_db()
        for sp in sheet_periods:
            if sp["period"]:
                cnt = conn_exists.execute("SELECT COUNT(*) FROM billing_area_readings WHERE period=?", (sp["period"],)).fetchone()[0]
                exists_map[sp["period"]] = cnt
        conn_exists.close()
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "header_row": header_row_idx if header_row_idx is not None else 0,
        "headers": raw_headers,
        "raw_rows": raw_rows,
        "parsed_rows": sample_parsed,
        "detected_period": detected_period,
        "sheet_periods": sheet_periods,
        "exists_map": exists_map,
        "mapping": field_map,
    })


@meters_bp.route("/api/billing/export")
def export_billing():
    import server
    period = request.args.get("period", "")
    fmt = request.args.get("format", "xlsx").lower()
    conn = server.get_db()
    if period:
        rows = conn.execute(
            "SELECT * FROM billing_area_readings WHERE period=? ORDER BY category, sort_order, id", (period,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM billing_area_readings ORDER BY period DESC, category, sort_order, id").fetchall()
    conn.close()

    def _fmt_period(p):
        try:
            y, m = p.split("-")
            return f"{int(y)}年{int(m)}月"
        except Exception:
            return p

    title = _fmt_period(period) + " 收费区域" if period else "全部月份 收费区域"

    if fmt == "csv":
        lines = ["\ufeff" + title]
        lines.append("名称,等级,地点,上月读数 KW.h,本月读数 KW.h,倍率,实际用电量 KW.h,月份")
        for r in rows:
            lines.append(",".join([
                str(r["name"] or ""),
                str(r["grade"] or ""),
                str(r["location"] or ""),
                str(r["prev_reading"] if r["prev_reading"] is not None else ""),
                str(r["curr_reading"] if r["curr_reading"] is not None else ""),
                str(r["multiplier"] if r["multiplier"] is not None else ""),
                str(r["usage_amount"] if r["usage_amount"] is not None else ""),
                str(r["period"] or ""),
            ]))
        fname = f"收费区域_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
        from urllib.parse import quote
        fname_encoded = quote(fname.encode("utf-8"), safe="")
        return Response("\n".join(lines), mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "收费区域"

    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["名称", "等级", "地点", "上月读数 KW.h", "本月读数 KW.h", "倍率", "实际用电量 KW.h"]
    if not period:
        headers.append("月份")
    ws.append(headers)
    header_row_idx = 2
    for c in ws[header_row_idx]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="DDEBF7")
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        row_data = [r["name"], r["grade"], r["location"],
                     r["prev_reading"], r["curr_reading"], r["multiplier"], r["usage_amount"]]
        if not period:
            row_data.append(r["period"])
        ws.append(row_data)

    widths = [24, 10, 16, 16, 16, 10, 18]
    if not period:
        widths.append(12)
    for i, w in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"收费区域_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from urllib.parse import quote
    fname_encoded = quote(fname.encode("utf-8"), safe="")
    return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})


@meters_bp.route("/api/cost/periods")
def cost_periods():
    import server
    conn = server.get_db()
    rows = conn.execute("SELECT DISTINCT period FROM cost_comparison ORDER BY period DESC").fetchall()
    conn.close()
    return jsonify([r["period"] for r in rows])


@meters_bp.route("/api/cost")
def list_cost():
    import server
    period = request.args.get("period", "")
    conn = server.get_db()
    if period:
        rows = conn.execute(
            "SELECT * FROM cost_comparison WHERE period=? ORDER BY id ASC", (period,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM cost_comparison ORDER BY period DESC, id ASC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@meters_bp.route("/api/cost", methods=["POST"])
def add_cost():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    conn = server.get_db()
    data = request.json or {}
    period = data.get("period", "")
    energy_type = data.get("energy_type", "")
    if not period or not energy_type:
        conn.close()
        return jsonify({"error": "月份和能耗类型必填"}), 400
    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn.execute("""INSERT INTO cost_comparison
        (period, energy_type, current_year, current_price, current_amount,
         previous_year, previous_price, previous_amount, comparison_rate, unit,
         created_by, updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
        (period, energy_type, data.get("current_year"), data.get("current_price"), data.get("current_amount"),
         data.get("previous_year"), data.get("previous_price"), data.get("previous_amount"),
         data.get("comparison_rate"), data.get("unit", ""),
         sess["username"] if sess else "",
         dt.now().isoformat()))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = conn.execute("SELECT * FROM cost_comparison WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@meters_bp.route("/api/cost/<int:cid>", methods=["PUT"])
def update_cost(cid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    conn = server.get_db()
    data = request.json or {}
    row = conn.execute("SELECT * FROM cost_comparison WHERE id=?", (cid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "记录不存在"}), 404
    conn.execute("""UPDATE cost_comparison SET
        period=?, energy_type=?, current_year=?, current_price=?, current_amount=?,
        previous_year=?, previous_price=?, previous_amount=?, comparison_rate=?, unit=?,
        updated_at=? WHERE id=?""",
        (data.get("period"), data.get("energy_type"), data.get("current_year"),
         data.get("current_price"), data.get("current_amount"),
         data.get("previous_year"), data.get("previous_price"), data.get("previous_amount"),
         data.get("comparison_rate"), data.get("unit", ""), dt.now().isoformat(), cid))
    conn.commit()
    row = conn.execute("SELECT * FROM cost_comparison WHERE id=?", (cid,)).fetchone()
    conn.close()
    return jsonify(dict(row))


@meters_bp.route("/api/cost/<int:cid>", methods=["DELETE"])
def delete_cost(cid):
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    conn = server.get_db()
    conn.execute("DELETE FROM cost_comparison WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@meters_bp.route("/api/cost/batch-delete", methods=["POST"])
def batch_delete_cost():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再操作"}), 403
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "请选择要删除的记录"}), 400
    conn = server.get_db()
    placeholders = ",".join("?" * len(ids))
    conn.execute(f"DELETE FROM cost_comparison WHERE id IN ({placeholders})", ids)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "deleted": len(ids)})


@meters_bp.route("/api/cost/import", methods=["POST"])
def import_cost():
    import server
    if not require_login_write(request): return jsonify({"error": "请先登录后台账号再导入"}), 403
    sess = server.get_session(request.headers.get("X-Session-Token", ""))
    conn = server.get_db()
    
    file = request.files.get("file")
    if not file:
        conn.close()
        return jsonify({"error": "请选择文件"}), 400
    
    period = request.form.get("period", "").strip()
    
    try:
        file_content = file.read()
        wb, err = _load_workbook(file_content, file.filename or "")
        if wb is None:
            conn.close()
            return jsonify({"error": err or "文件格式不支持，请上传 .xls / .xlsx / .csv 文件"}), 400
    except Exception as e:
        conn.close()
        return jsonify({"error": f"文件解析失败: {e}"}), 400
    
    import re
    
    def safe_float(v):
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").replace("，", "").strip())
        except (ValueError, TypeError):
            return None
    
    def parse_period_from_name(name):
        m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", name)
        if m:
            y = int(m.group(1))
            mo = int(m.group(2))
            if 1900 <= y <= 2100 and 1 <= mo <= 12:
                return f"{y:04d}-{mo:02d}"
        return None
    
    all_entries = []
    
    for ws in wb.worksheets:
        sheet_name = str(ws.title).strip()
        if not sheet_name.startswith("能耗"):
            continue
        
        sheet_period = parse_period_from_name(sheet_name)
        if not sheet_period:
            continue
        
        if period and period != sheet_period:
            continue
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < 2:
                continue
            
            energy_type = str(row[0]).strip() if row[0] else ""
            if not energy_type:
                continue
            
            row_text = "".join(str(v).strip() for v in row if v)
            skip_keywords = ("合计", "总表", "抄表日", "制表人：", "制表人:", "审核", "负责人", "单位", "名称", "项目")
            if any(kw in row_text for kw in skip_keywords):
                continue
            
            has_numeric = any(safe_float(row[i]) is not None for i in range(1, min(8, len(row))))
            if not has_numeric:
                continue
            
            unit = ""
            if "/" in energy_type:
                parts = energy_type.split("/")
                energy_type = parts[0].strip()
                unit = parts[1].strip()
            
            all_entries.append({
                "period": sheet_period,
                "energy_type": energy_type,
                "unit": unit,
                "current_year": safe_float(row[1]),
                "current_price": safe_float(row[2]),
                "current_amount": safe_float(row[3]),
                "previous_year": safe_float(row[4]),
                "previous_price": safe_float(row[5]),
                "previous_amount": safe_float(row[6]),
                "comparison_rate": safe_float(row[7]),
            })
    
    if not all_entries:
        conn.close()
        if period:
            return jsonify({"error": f"未找到月份为 {period} 的能耗工作表"}), 400
        else:
            return jsonify({"error": "未能识别任何有效的能耗工作表，请检查文件内容"}), 400
    
    imported_by_period = {}
    for e in all_entries:
        p = e["period"]
        if p not in imported_by_period:
            imported_by_period[p] = 0
        imported_by_period[p] += 1
    
    username = sess["username"] if sess else ""
    now = dt.now().isoformat()
    
    for p in imported_by_period:
        existing = conn.execute("SELECT COUNT(*) FROM cost_comparison WHERE period=?", (p,)).fetchone()[0]
        conn.execute("DELETE FROM cost_comparison WHERE period=?", (p,))
        
        period_entries = [e for e in all_entries if e["period"] == p]
        rows = [(p, e["energy_type"], e["current_year"], e["current_price"], e["current_amount"],
                 e["previous_year"], e["previous_price"], e["previous_amount"], e["comparison_rate"], e["unit"],
                 username, now) for e in period_entries]
        conn.executemany("""INSERT INTO cost_comparison
            (period, energy_type, current_year, current_price, current_amount,
             previous_year, previous_price, previous_amount, comparison_rate, unit,
             created_by, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
    
    conn.commit()
    conn.close()
    
    result = {"ok": True, "imported": len(all_entries), "periods": imported_by_period}
    return jsonify(result)


@meters_bp.route("/api/cost/export")
def export_cost():
    import server
    period = request.args.get("period", "")
    fmt = request.args.get("format", "xlsx").lower()
    conn = server.get_db()
    if period:
        rows = conn.execute(
            "SELECT * FROM cost_comparison WHERE period=? ORDER BY id ASC", (period,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM cost_comparison ORDER BY period DESC, id ASC").fetchall()
    conn.close()

    def _fmt_period(p):
        try:
            y, m = p.split("-")
            return f"{int(y)}年{int(m)}月"
        except Exception:
            return p

    if period:
        title = f"浙江出版物资大厦有限公司月能耗对比表（{_fmt_period(period)}）"
    else:
        title = "浙江出版物资大厦有限公司月能耗对比表"

    if fmt == "csv":
        lines = ["\ufeff" + title]
        lines.append("能耗,本年数量,本年单价,本年金额,上年数量,上年单价,上年金额,对比率")
        if not period:
            lines[-1] += ",月份"
        for r in rows:
            energy_type = str(r["energy_type"] or "")
            if r["unit"]:
                energy_type += " /" + str(r["unit"])
            line_data = [
                energy_type,
                str(r["current_year"] if r["current_year"] is not None else ""),
                str(r["current_price"] if r["current_price"] is not None else ""),
                str(r["current_amount"] if r["current_amount"] is not None else ""),
                str(r["previous_year"] if r["previous_year"] is not None else ""),
                str(r["previous_price"] if r["previous_price"] is not None else ""),
                str(r["previous_amount"] if r["previous_amount"] is not None else ""),
                str(round(r["comparison_rate"] * 100, 2)) + "%" if r["comparison_rate"] is not None else "",
            ]
            if not period:
                line_data.append(str(r["period"] or ""))
            lines.append(",".join(line_data))
        fname = f"费用对比_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
        from urllib.parse import quote
        fname_encoded = quote(fname.encode("utf-8"), safe="")
        return Response("\n".join(lines), mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "能耗对比"

    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["能耗", "本年数量", "本年单价", "本年金额", "上年数量", "上年单价", "上年金额", "对比率"]
    if not period:
        headers.append("月份")
    ws.append(headers)
    header_row_idx = 2
    for c in ws[header_row_idx]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="DDEBF7")
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        energy_type = str(r["energy_type"] or "")
        if r["unit"]:
            energy_type += " /" + str(r["unit"])
        row_data = [energy_type,
                     r["current_year"], r["current_price"], r["current_amount"],
                     r["previous_year"], r["previous_price"], r["previous_amount"],
                     round(r["comparison_rate"] * 100, 2) if r["comparison_rate"] is not None else None]
        if not period:
            row_data.append(r["period"])
        ws.append(row_data)

    widths = [16, 8, 12, 12, 12, 12, 12, 12, 10]
    if not period:
        widths.append(12)
    for i, w in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"费用对比_{period or 'all'}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from urllib.parse import quote
    fname_encoded = quote(fname.encode("utf-8"), safe="")
    return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename*=UTF-8''{fname_encoded}"})